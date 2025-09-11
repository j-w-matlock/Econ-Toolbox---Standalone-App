import copy
import streamlit as st
import numpy as np
import pandas as pd
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference

# Conversion table from point rankings to unit day values ($/user day)
POINT_VALUE_TABLE = pd.DataFrame(
    {
        "Points": [0, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100],
        "General Recreation": [
            4.87,
            5.78,
            6.39,
            7.31,
            9.13,
            10.35,
            11.26,
            11.87,
            13.09,
            14.00,
            14.61,
        ],
        "General Fishing and Hunting": [
            7.00,
            7.91,
            8.52,
            9.44,
            10.35,
            11.26,
            12.48,
            13.09,
            14.00,
            14.31,
            14.61,
        ],
        "Specialized Fishing and Hunting": [
            34.09,
            35.01,
            35.62,
            36.53,
            37.44,
            41.10,
            44.75,
            47.49,
            51.14,
            54.80,
            57.84,
        ],
        "Specialized Recreation": [
            19.79,
            21.00,
            22.53,
            24.35,
            25.88,
            29.22,
            32.27,
            38.97,
            45.36,
            51.75,
            57.84,
        ],
    }
)


st.title("Economic Toolbox")


def persistent_data_editor(df: pd.DataFrame, key: str, **kwargs) -> pd.DataFrame:
    """Return an edited copy of ``df`` that preserves first changes.

    ``st.data_editor`` mutates its input in place and may return a view of the
    original object, causing the first user edit to disappear on rerun. Passing
    a deep copy to the widget and returning a deep copy of the edited result
    ensures that all edits persist immediately.
    """

    edited = st.data_editor(copy.deepcopy(df), key=key, **kwargs)
    return copy.deepcopy(edited)


def ead_trapezoidal(prob, damages):
    """Return expected annual damage via trapezoidal integration."""
    prob = np.asarray(prob, dtype=float)
    damages = np.asarray(damages, dtype=float)
    return float(
        np.sum(0.5 * (damages[:-1] + damages[1:]) * (prob[:-1] - prob[1:]))
    )


def updated_storage_cost(tc, sp, storage_reallocated, total_usable_storage):
    """Compute updated cost of storage for reservoir reallocations."""
    tc = float(tc)
    sp = float(sp)
    storage_reallocated = float(storage_reallocated)
    total_usable_storage = float(total_usable_storage)
    return (tc - sp) * storage_reallocated / total_usable_storage


def interest_during_construction(
    total_initial_cost,
    rate,
    months,
    *,
    costs=None,
    timings=None,
    normalize=True,
):
    """Compute interest during construction (IDC).

    Parameters
    ----------
    total_initial_cost : float
        Total initial cost excluding IDC. Used when ``costs`` is not provided.
    rate : float
        Annual interest rate expressed as a decimal (e.g., ``0.05`` for 5%).
    months : int
        Construction period in months.
    costs : list[float], optional
        Explicit costs for each month. If provided, ``normalize`` is ignored.
    timings : list[str], optional
        Timing of each cost within the month: ``"beginning"``, ``"middle"``, or
        ``"end"``. Defaults to ``"middle"`` for any unspecified timing.
    normalize : bool, default ``True``
        When ``True`` and ``costs`` is ``None``, the ``total_initial_cost`` is
        distributed evenly across all months with the first month treated as a
        beginning-of-month expenditure and remaining months at midpoints.

    Returns
    -------
    float
        Interest accrued during construction.
    """
    if months <= 0:
        return 0.0

    monthly_rate = rate / 12.0

    if costs is None:
        if not normalize:
            # Legacy approximation assuming evenly spread costs.
            years = months / 12.0
            return total_initial_cost * rate * years / 8

        monthly_cost = total_initial_cost / months
        costs = [monthly_cost] * months
        timings = ["beginning"] + ["middle"] * (months - 1)
    else:
        if timings is None:
            timings = ["middle"] * len(costs)

    idc = 0.0
    for i, cost in enumerate(costs, start=1):
        timing = timings[i - 1]
        if timing == "beginning":
            remaining = months - i + 1
        elif timing == "end":
            remaining = months - i
        else:  # middle
            remaining = months - i + 0.5
        idc += cost * monthly_rate * remaining
    return idc


def capital_recovery_factor(rate, periods):
    """Return capital recovery factor for a given discount rate and period."""
    if rate == 0:
        return 1 / periods
    return rate * (1 + rate) ** periods / ((1 + rate) ** periods - 1)


def build_excel():
    """Assemble all inputs, results, and README into an Excel workbook."""
    buffer = BytesIO()
    wb = Workbook()

    # EAD inputs sheet
    ws_ead_inputs = wb.active
    ws_ead_inputs.title = "EAD Inputs"
    table = st.session_state.get("table")
    if isinstance(table, pd.DataFrame):
        for row in dataframe_to_rows(table, index=False, header=True):
            ws_ead_inputs.append(row)

    # Charts sheet
    charts_for_export = st.session_state.get("charts_for_export", [])
    if charts_for_export:
        ws_charts = wb.create_sheet("Charts")
        for chart_info in charts_for_export:
            df_chart = chart_info["data"]
            start_row = ws_charts.max_row + 2 if ws_charts.max_row > 1 else 1
            for row in dataframe_to_rows(df_chart, index=False, header=True):
                ws_charts.append(row)
            end_row = start_row + len(df_chart)
            chart = LineChart()
            chart.title = chart_info["title"]
            chart.y_axis.title = df_chart.columns[1]
            chart.x_axis.title = df_chart.columns[0]
            data_ref = Reference(
                ws_charts, min_col=2, min_row=start_row, max_row=end_row
            )
            chart.add_data(data_ref, titles_from_data=True)
            cats_ref = Reference(
                ws_charts, min_col=1, min_row=start_row + 1, max_row=end_row
            )
            chart.set_categories(cats_ref)
            ws_charts.add_chart(chart, f"E{start_row}")

    # EAD results
    if st.session_state.get("ead_results"):
        ws_ead = wb.create_sheet("EAD Results")
        for k, v in st.session_state.ead_results.items():
            ws_ead.append([k, v])
        diffs = st.session_state.get("ead_differences", {})
        pct_changes = st.session_state.get("ead_percent_changes", {})
        if diffs:
            ws_ead.append([])
            for k, v in diffs.items():
                ws_ead.append([f"{k} - Damage 1", v])
        if pct_changes:
            ws_ead.append([])
            for k, v in pct_changes.items():
                ws_ead.append([f"{k} % change from Damage 1", v])

    # Storage capacity inputs and results
    storage_capacity = st.session_state.get("storage_capacity")
    if storage_capacity:
        ws_sc = wb.create_sheet("Storage Capacity")
        ws_sc.append(["Total Usable Storage (STot)", storage_capacity.get("STot")])
        ws_sc.append(["Storage Recommendation (SRec)", storage_capacity.get("SRec")])
        ws_sc.append(
            [
                "Percent of Total Conservation Storage (P)",
                storage_capacity.get("P"),
            ]
        )

    # Joint costs O&M inputs and results
    joint_om = st.session_state.get("joint_om")
    if joint_om:
        ws_jom = wb.create_sheet("Joint Costs O&M")
        ws_jom.append(["Joint Operations Cost ($/year)", joint_om.get("operations")])
        ws_jom.append(
            ["Joint Maintenance Cost ($/year)", joint_om.get("maintenance")]
        )
        ws_jom.append(["Total Joint O&M", joint_om.get("total")])

    # Updated storage costs table and summary
    updated_storage = st.session_state.get("updated_storage")
    if updated_storage:
        ws_usc = wb.create_sheet("Updated Storage Costs")
        table = updated_storage.get("table", pd.DataFrame())
        if isinstance(table, pd.DataFrame):
            for row in dataframe_to_rows(table, index=False, header=True):
                ws_usc.append(row)
        if "CTot" in updated_storage:
            ws_usc.append([])
            ws_usc.append([
                "Total Updated Cost of Storage",
                updated_storage.get("CTot"),
            ])

    # RR&R and Mitigation inputs and results
    rrr_mit = st.session_state.get("rrr_mit")
    if rrr_mit:
        ws_rrr = wb.create_sheet("RR&R and Mitigation")
        ws_rrr.append(["Federal Discount Rate (%)", rrr_mit.get("rate")])
        ws_rrr.append(["Analysis Years (Periods)", rrr_mit.get("periods")])
        ws_rrr.append(["CWCCI Ratio", rrr_mit.get("cwcci")])
        ws_rrr.append(["Base Year", rrr_mit.get("base_year")])
        table = rrr_mit.get("table")
        if isinstance(table, pd.DataFrame) and not table.empty:
            ws_rrr.append([])
            for row in dataframe_to_rows(table, index=False, header=True):
                ws_rrr.append(row)
            ws_rrr.append([])
        ws_rrr.append(["Total Present Value Cost", rrr_mit.get("total_pv")])
        ws_rrr.append(["Updated Cost", rrr_mit.get("updated_cost")])
        ws_rrr.append(
            [
                "Annualized RR&R and Mitigation",
                rrr_mit.get("annualized"),
            ]
        )

    # Total annual cost summary
    total_inputs = st.session_state.get("total_annual_cost_inputs")
    storage_costs = st.session_state.get("storage_cost")
    if total_inputs or storage_costs is not None:
        ws_tac = wb.create_sheet("Total Annual Cost")
        p = storage_capacity.get("P") if storage_capacity else None
        ctot = updated_storage.get("CTot") if updated_storage else None
        om_total = joint_om.get("total") if joint_om else None
        rrr_annual = rrr_mit.get("annualized") if rrr_mit else None
        om_scaled = om_total * p if None not in (om_total, p) else None
        rrr_scaled1 = rrr_annual * p if None not in (rrr_annual, p) else None

        drate1 = total_inputs.get("rate1") if total_inputs else None
        years1 = total_inputs.get("periods1") if total_inputs else None
        drate2 = total_inputs.get("rate2") if total_inputs else None
        years2 = total_inputs.get("periods2") if total_inputs else None

        cap1 = cap2 = None
        crec = None
        if None not in (ctot, p):
            crec = ctot * p
        if None not in (ctot, p, drate1, years1):
            cap1 = ctot * p * capital_recovery_factor(drate1 / 100.0, years1)
        if None not in (ctot, p, drate2, years2):
            cap2 = ctot * p * capital_recovery_factor(drate2 / 100.0, years2)

        ws_tac.append(["Metric", "Scenario 1", "Scenario 2"])
        if p is not None:
            ws_tac.append(["Percent of Total Conservation Storage (P)", p, p])
        if crec is not None:
            ws_tac.append(["Cost of Storage Recommendation", crec, crec])
        if cap1 is not None or cap2 is not None:
            ws_tac.append(["Annualized Storage Cost", cap1, cap2])
        if om_scaled is not None:
            ws_tac.append(["Joint O&M", om_scaled, om_scaled])
        if rrr_scaled1 is not None:
            ws_tac.append(["Annualized RR&R/Mitigation", rrr_scaled1, 0.0])
        if isinstance(storage_costs, dict):
            ws_tac.append(
                [
                    "Total Annual Cost",
                    storage_costs.get("scenario1"),
                    storage_costs.get("scenario2"),
                ]
            )
        if drate1 is not None or drate2 is not None:
            ws_tac.append(["Discount Rate (%) for Storage Cost", drate1, drate2])
        if years1 is not None or years2 is not None:
            ws_tac.append(["Analysis Period (years)", years1, years2])

    # Annualizer inputs, future costs, and summary
    if (
        st.session_state.get("annualizer_inputs")
        or st.session_state.get("future_costs_df") is not None
        or st.session_state.get("annualizer_summary")
    ):
        ws_ann = wb.create_sheet("Annualizer")
        for k, v in st.session_state.get("annualizer_inputs", {}).items():
            ws_ann.append([k, v])
        if st.session_state.get("annualizer_inputs"):
            ws_ann.append([])
        future_df = st.session_state.get("future_costs_df", pd.DataFrame())
        if not future_df.empty:
            for row in dataframe_to_rows(future_df, index=False, header=True):
                ws_ann.append(row)
            ws_ann.append([])
        for k, v in st.session_state.get("annualizer_summary", {}).items():
            ws_ann.append([k, v])

    # UDV analysis inputs and result
    if (
        st.session_state.get("udv_inputs")
        or st.session_state.get("udv_benefit")
    ):
        ws_rec = wb.create_sheet("UDV Analysis")
        for k, v in st.session_state.get("udv_inputs", {}).items():
            ws_rec.append([k, v])
        if "udv_benefit" in st.session_state:
            ws_rec.append(["Annual Recreation Benefit", st.session_state.udv_benefit])

    # Water demand forecast results
    water_df = st.session_state.get("water_demand_results")
    if isinstance(water_df, pd.DataFrame) and not water_df.empty:
        ws_water = wb.create_sheet("Water Demand")
        for row in dataframe_to_rows(water_df, index=False, header=True):
            ws_water.append(row)

    # README sheet
    readme_lines = Path("README.md").read_text().splitlines()
    ws_readme = wb.create_sheet("README")
    for line in readme_lines:
        ws_readme.append([line])

    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_button():
    """Render a download button for the current workbook."""
    buffer = build_excel()
    st.download_button(
        label="Export to Excel",
        data=buffer,
        file_name="econ_toolbox.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Export all inputs, results, and the README as an Excel file.",
    )


def ead_calculator():
    """EAD calculator with data entry, charts, and results."""
    st.write(
        "Compute expected annual damages using the U.S. Army Corps of Engineers trapezoidal method."
    )
    st.caption(
        "Reference: U.S. Army Corps of Engineers, Engineering Manual 1110-2-1619 (1996)."
    )

    st.subheader("Inputs")
    st.info(
        "Fill in the frequency and damage values below. Use the checkbox to add a stage column and the button to insert additional damage columns as needed."
    )

    if "num_damage_cols" not in st.session_state:
        st.session_state.num_damage_cols = 1

    if "table" not in st.session_state:
        st.session_state.table = pd.DataFrame(
            {
                "Frequency": [1.00, 0.50, 0.20, 0.10, 0.04, 0.02, 0.01, 0.005, 0.002],
                "Damage 1": [
                    0,
                    10000,
                    40000,
                    80000,
                    120000,
                    160000,
                    200000,
                    250000,
                    300000,
                ],
            }
        )

    include_stage = st.checkbox(
        "Include stage column",
        value="Stage" in st.session_state.table.columns,
        help="Add a column for stage values to enable stage-related charts.",
    )
    if include_stage and "Stage" not in st.session_state.table.columns:
        st.session_state.table.insert(1, "Stage", [None] * len(st.session_state.table))
    elif not include_stage and "Stage" in st.session_state.table.columns:
        st.session_state.table.drop(columns="Stage", inplace=True)

    if st.button(
        "Add damage column",
        help="Insert another damage column to compare scenarios.",
    ):
        st.session_state.num_damage_cols += 1
        st.session_state.table[f"Damage {st.session_state.num_damage_cols}"] = [
            None
        ] * len(st.session_state.table)

    column_config = {
        "Frequency": st.column_config.NumberColumn(
            "Frequency", min_value=0, max_value=1, step=0.001
        )
    }
    for col in st.session_state.table.columns:
        if col.startswith("Damage"):
            column_config[col] = st.column_config.NumberColumn(
                col, min_value=0, step=1000, format="$%d"
            )

    with st.form("data_table_form"):
        data = persistent_data_editor(
            st.session_state.table,
            key="table_editor",
            num_rows="dynamic",
            width="stretch",
            column_config=column_config,
        )
        submitted = st.form_submit_button(
            "Save table", help="Apply edits to the table above."
        )
    if submitted:
        # Store a separate copy so the saved table reflects the user's edits
        # even after further interaction.
        st.session_state.table = data.copy()

    damage_cols = [c for c in st.session_state.table.columns if c.startswith("Damage")]
    charts_for_export = []
    chart_data = (
        st.session_state.table.dropna(subset=["Frequency"])
        .sort_values("Frequency")
        .set_index("Frequency")[damage_cols]
    )
    if not chart_data.empty and damage_cols:
        st.subheader("Damage-Frequency Curve")
        selected_damage = st.selectbox(
            "Select damage column",
            damage_cols,
            key="df_damage",
            help="Choose which damage column to visualize.",
        )
        st.line_chart(chart_data[[selected_damage]])
        charts_for_export.append(
            {
                "title": "Damage-Frequency Curve",
                "data": chart_data[[selected_damage]].reset_index(),
            }
        )

    if "Stage" in st.session_state.table.columns:
        stage_df = (
            st.session_state.table.dropna(subset=["Stage"])
            .assign(Stage=lambda df: pd.to_numeric(df["Stage"], errors="coerce"))
            .dropna(subset=["Stage"])
            .set_index("Stage")
        )
        if not stage_df.empty:
            if damage_cols:
                st.subheader("Stage-Damage Curve")
                dmg_col = st.selectbox(
                    "Select damage column (stage)",
                    damage_cols,
                    key="stage_damage",
                    help="Damage column to plot against stage values.",
                )
                st.line_chart(stage_df[[dmg_col]])
                charts_for_export.append(
                    {
                        "title": "Stage-Damage Curve",
                        "data": stage_df[[dmg_col]].reset_index(),
                    }
                )
            if "Frequency" in stage_df.columns:
                st.subheader("Stage-Frequency Curve")
                st.line_chart(stage_df["Frequency"])
                charts_for_export.append(
                    {
                        "title": "Stage-Frequency Curve",
                        "data": stage_df[["Frequency"]].reset_index(),
                    }
                )

    st.session_state.charts_for_export = charts_for_export

    st.subheader("EAD Results")
    st.info(
        "Click the button below to compute expected annual damages for each damage column using trapezoidal integration."
    )
    if st.button(
        "Calculate EAD",
        help="Run the trapezoidal method on the frequency and damage data.",
    ):
        df = st.session_state.table.dropna(subset=["Frequency"]).sort_values(
            "Frequency", ascending=False
        )
        freq = df["Frequency"].to_numpy()
        if not np.isclose(freq[0], 1.0) or np.any(np.diff(freq) > 0):
            st.warning("Frequencies should start at 1 and monotonically decrease to 0.")
        missing_zero = freq[-1] != 0
        if missing_zero:
            st.info(
                "Final frequency not 0; appending zero-frequency point using last damage value."
            )
        results = {}
        for col in df.columns:
            if col.startswith("Damage"):
                damages = df[col].fillna(0).to_numpy()
                freq_use = np.append(freq, 0.0) if missing_zero else freq
                damages_use = (
                    np.append(damages, damages[-1]) if missing_zero else damages
                )
                if len(freq_use) >= 2 and len(freq_use) == len(damages_use):
                    results[col] = ead_trapezoidal(freq_use, damages_use)
                else:
                    results[col] = None
        if not results:
            st.error("No damage columns found.")
        else:
            base_ead = results.get("Damage 1")
            differences = {}
            pct_changes = {}
            for col, val in results.items():
                if val is None:
                    st.error(
                        f"{col}: Ensure at least two paired frequency and damage values."
                    )
                else:
                    st.success(f"{col} Expected Annual Damage: ${val:,.2f}")
                    if col != "Damage 1" and base_ead is not None:
                        diff = val - base_ead
                        differences[col] = diff
                        pct = (diff / base_ead * 100) if base_ead != 0 else np.nan
                        pct_changes[col] = pct
                        sign = "+" if diff >= 0 else "-"
                        st.info(
                            f"Difference from Damage 1: {sign}${abs(diff):,.2f} ({pct:+.2f}%)"
                        )
            st.session_state.ead_results = results
            st.session_state.ead_differences = differences
            st.session_state.ead_percent_changes = pct_changes

    export_button()


def storage_calculator():
    """Multi-step storage cost and O&M workflow."""
    st.header("Storage Cost and O&M Calculator")
    st.caption(
        "Replicates the workflow of the 'Updated Cost of Storage' spreadsheet."
    )

    tabs = st.tabs(
        [
            "Storage Capacity",
            "Joint Costs O&M",
            "Updated Storage Costs",
            "RR&R and Mitigation",
            "Total Annual Cost",
        ]
    )

    # --- Storage Capacity -------------------------------------------------
    with tabs[0]:
        st.subheader("Storage Capacity")
        st.info(
            "Determine the percent of conservation storage recommended for water supply."
        )
        st.session_state.setdefault("storage_capacity", {})
        stot = st.number_input(
            "Total Usable Storage (STot) (ac-ft)",
            min_value=0.0,
            value=float(st.session_state.storage_capacity.get("STot", 0.0)),
            help="Total usable conservation storage (cell B2).",
        )
        srec = st.number_input(
            "Storage Recommendation (SRec) (ac-ft)",
            min_value=0.0,
            value=float(st.session_state.storage_capacity.get("SRec", 0.0)),
            help="Storage volume proposed for reallocation (cell B3).",
        )
        p = (srec / stot) if stot else 0.0
        st.write(f"Percent of Total Conservation Storage (P): {p:.5f}")
        st.session_state.storage_capacity = {"STot": stot, "SRec": srec, "P": p}

    # --- Joint Costs O&M --------------------------------------------------
    with tabs[1]:
        st.subheader("Joint Costs O&M")
        st.info(
            "Enter annual joint operations and maintenance costs updated to the current price level."
        )
        st.session_state.setdefault("joint_om", {})
        ops = st.number_input(
            "Joint Operations Cost ($/year)",
            min_value=0.0,
            value=float(st.session_state.joint_om.get("operations", 0.0)),
            help="Sum of joint operations expenditures.",
        )
        maint = st.number_input(
            "Joint Maintenance Cost ($/year)",
            min_value=0.0,
            value=float(st.session_state.joint_om.get("maintenance", 0.0)),
            help="Sum of joint maintenance expenditures.",
        )
        total_om = ops + maint
        st.write(f"Total Joint O&M: ${total_om:,.2f}")
        st.session_state.joint_om = {
            "operations": ops,
            "maintenance": maint,
            "total": total_om,
        }

    # --- Updated Storage Costs -------------------------------------------
    with tabs[2]:
        st.subheader("Updated Storage Costs")
        st.info(
            "Update original joint-use construction costs to current dollars using CWCCIS ratios."
        )
        if "usc_table" not in st.session_state:
            st.session_state.usc_table = pd.DataFrame(
                {
                    "Category": ["Lands and Damages", "Relocations", "Dam"],
                    "Actual Cost": [0.0, 0.0, 0.0],
                    "Update Factor": [1.0, 1.0, 1.0],
                }
            )
        usc_cols = {
            "Actual Cost": st.column_config.NumberColumn(
                "Actual Cost", min_value=0.0, format="$%.2f"
            ),
            "Update Factor": st.column_config.NumberColumn(
                "Update Factor", min_value=0.0, format="%.5f"
            ),
        }
        raw_table = persistent_data_editor(
            st.session_state.usc_table,
            key="usc_table_editor",
            num_rows="dynamic",
            width="stretch",
            column_config=usc_cols,
        )
        st.session_state.usc_table = raw_table.copy()
        table = raw_table.assign(
            **{"Updated Cost": raw_table["Actual Cost"] * raw_table["Update Factor"]}
        )
        st.table(table)
        ctot = float(table["Updated Cost"].sum())
        st.write(f"Total Updated Cost of Storage (CTot): ${ctot:,.2f}")
        st.session_state.updated_storage = {"table": table, "CTot": ctot}

    # --- RR&R and Mitigation ---------------------------------------------
    with tabs[3]:
        st.subheader("RR&R and Mitigation")
        st.info(
            "Annualize rehabilitation/replacement and mitigation costs using a capital recovery factor."
        )
        st.session_state.setdefault("rrr_mit", {})
        rate = st.number_input(
            "Federal Discount Rate (%)",
            min_value=0.0,
            value=float(st.session_state.rrr_mit.get("rate", 0.0)),
            help="Cell B2: federal discount rate expressed as a percent.",
        )
        periods = st.number_input(
            "Analysis Years (Periods)",
            min_value=1,
            step=1,
            value=int(st.session_state.rrr_mit.get("periods", 30)),
            help="Cell B3: number of years over which to annualize costs.",
        )
        cwcci = st.number_input(
            "CWCCI Ratio (FY/FY)",
            min_value=0.0,
            value=float(st.session_state.rrr_mit.get("cwcci", 1.0)),
            help="Cell B4: ratio of CWCCIS indices to update costs to the base year.",
        )
        base_year = st.number_input(
            "Base Year",
            min_value=0,
            step=1,
            value=int(st.session_state.rrr_mit.get("base_year", 0)),
            help="Year to which future costs are discounted.",
        )
        if "rrr_costs" not in st.session_state:
            # Use explicit dtypes so Streamlit treats the ``Item`` column as text.
            st.session_state.rrr_costs = pd.DataFrame(
                {
                    "Item": pd.Series(dtype="object"),
                    "Future Cost": pd.Series(dtype="float"),
                    "Year": pd.Series(dtype="int"),
                }
            )
        cost_cols = {
            "Item": st.column_config.TextColumn("Item"),
            "Future Cost": st.column_config.NumberColumn(
                "Future Cost", min_value=0.0, format="$%.2f"
            ),
            "Year": st.column_config.NumberColumn(
                "Year", min_value=0, step=1, format="%d"
            ),
        }
        raw_costs = persistent_data_editor(
            st.session_state.rrr_costs,
            key="rrr_costs_editor",
            num_rows="dynamic",
            width="stretch",
            column_config=cost_cols,
        )
        # Preserve dtypes (especially ``Item``) on update.
        st.session_state.rrr_costs = (
            raw_costs.astype({"Item": "object", "Future Cost": "float", "Year": "int"}, errors="ignore").copy()
        )
        rate_dec = rate / 100.0
        if not raw_costs.empty:
            costs = raw_costs.copy()
            costs["PV Factor"] = 1 / (
                (1 + rate_dec) ** (costs["Year"] - base_year)
            )
            costs["Present Value"] = costs["Future Cost"] * costs["PV Factor"]
            st.table(costs)
            total_pv = float(costs["Present Value"].sum())
        else:
            costs = pd.DataFrame()
            total_pv = 0.0
        updated_cost = total_pv * cwcci
        crf = capital_recovery_factor(rate_dec, periods)
        annualized = updated_cost * crf
        st.write(f"Updated Cost: ${updated_cost:,.2f}")
        st.write(f"Annualized RR&R and Mitigation: ${annualized:,.2f}")
        st.session_state.rrr_mit = {
            "rate": rate,
            "periods": periods,
            "cwcci": cwcci,
            "base_year": base_year,
            "table": costs,
            "total_pv": total_pv,
            "updated_cost": updated_cost,
            "annualized": annualized,
        }

    # --- Total Annual Cost ------------------------------------------------
    with tabs[4]:
        st.subheader("Total Annual Cost")
        st.info(
            "Combine capital, O&M, and RR&R/mitigation to estimate the annual cost of reallocation."
        )
        # Inputs shared across scenarios
        p = st.session_state.get("storage_capacity", {}).get("P", 0.0)
        ctot = st.session_state.get("updated_storage", {}).get("CTot", 0.0)
        om_total = st.session_state.get("joint_om", {}).get("total", 0.0)
        rrr_annual = st.session_state.get("rrr_mit", {}).get("annualized", 0.0)
        om_scaled = om_total * p
        rrr_scaled = rrr_annual * p
        rrr_scaled2 = 0.0
        crec = ctot * p

        st.session_state.setdefault("total_annual_cost_inputs", {})
        inputs = st.session_state.total_annual_cost_inputs

        col1, col2 = st.columns(2)

        with col1:
            drate1 = st.number_input(
                "Discount Rate (%)",
                min_value=0.0,
                value=float(
                    inputs.get("rate1", st.session_state.rrr_mit.get("rate", 0.0))
                ),
                key="tac_rate1",
                help="Discount rate used to annualize updated storage costs.",
            )
            years1 = st.number_input(
                "Analysis Period (years)",
                min_value=1,
                step=1,
                value=int(inputs.get("periods1", 30)),
                key="tac_years1",
                help="Number of years over which storage costs are annualized.",
            )
            capital1 = ctot * p * capital_recovery_factor(drate1 / 100.0, years1)
            total1 = capital1 + om_scaled + rrr_scaled
            st.metric("Percent of Total Conservation Storage (P)", f"{p:.5f}")
            st.metric("Cost of Storage Recommendation", f"${crec:,.2f}")
            st.metric("Annualized Storage Cost", f"${capital1:,.2f}")
            st.metric("Joint O&M", f"${om_scaled:,.2f}")
            st.metric("Annualized RR&R/Mitigation", f"${rrr_scaled:,.2f}")
            st.metric("Total Annual Cost", f"${total1:,.2f}")

        with col2:
            drate2 = st.number_input(
                "Discount Rate (%)",
                min_value=0.0,
                value=float(
                    inputs.get("rate2", st.session_state.rrr_mit.get("rate", 0.0))
                ),
                key="tac_rate2",
                help="Discount rate used to annualize updated storage costs.",
            )
            years2 = st.number_input(
                "Analysis Period (years)",
                min_value=1,
                step=1,
                value=int(inputs.get("periods2", 50)),
                key="tac_years2",
                help="Number of years over which storage costs are annualized.",
            )
            capital2 = ctot * p * capital_recovery_factor(drate2 / 100.0, years2)
            total2 = capital2 + om_scaled
            st.metric("Percent of Total Conservation Storage (P)", f"{p:.5f}")
            st.metric("Cost of Storage Recommendation", f"${crec:,.2f}")
            st.metric("Annualized Storage Cost", f"${capital2:,.2f}")
            st.metric("Joint O&M", f"${om_scaled:,.2f}")
            st.metric("Annualized RR&R/Mitigation", f"${rrr_scaled2:,.2f}")
            st.metric("Total Annual Cost", f"${total2:,.2f}")

        st.session_state.total_annual_cost_inputs = {
            "rate1": drate1,
            "periods1": years1,
            "rate2": drate2,
            "periods2": years2,
        }
        st.session_state.storage_inputs = {
            **st.session_state.get("storage_capacity", {}),
            **st.session_state.get("joint_om", {}),
            **st.session_state.get("rrr_mit", {}),
            "Capital Discount Rate 1": drate1,
            "Capital Analysis Years 1": years1,
            "Capital Discount Rate 2": drate2,
            "Capital Analysis Years 2": years2,
            "Updated Cost Total": ctot,
        }
        st.session_state.storage_cost = {
            "scenario1": total1,
            "scenario2": total2,
        }

    export_button()


def annualizer_calculator():
    """Project cost annualizer."""
    st.header("Project Cost Annualizer")
    st.info("Calculate annualized project costs and benefit-cost ratio.")

    if "num_future_costs" not in st.session_state:
        st.session_state.num_future_costs = 0

    if st.button("Add planned future cost"):
        st.session_state.num_future_costs += 1

    with st.form("annualizer_form"):
        first_cost = st.number_input(
            "Project First Cost ($)", min_value=0.0, value=0.0
        )
        real_estate_cost = st.number_input(
            "Real Estate Cost ($)", min_value=0.0, value=0.0
        )
        ped_cost = st.number_input("PED Cost ($)", min_value=0.0, value=0.0)
        monitoring_cost = st.number_input(
            "Monitoring Cost ($)", min_value=0.0, value=0.0
        )
        idc_rate = st.number_input(
            "Interest Rate (%) - For Interest During Construction",
            min_value=0.0,
            value=0.0,
        )
        construction_months = st.number_input(
            "Construction Period (Months)", min_value=0, value=0, step=1
        )
        idc_method = st.radio(
            "IDC Cost Distribution",
            [
                "Normalize over construction period",
                "Specify per-period costs",
            ],
            index=0,
        )
        period_costs = []
        period_timings = []
        if idc_method == "Specify per-period costs":
            for m in range(1, int(construction_months) + 1):
                c = st.number_input(
                    f"Cost in month {m} ($)",
                    min_value=0.0,
                    value=0.0,
                    key=f"month_cost_{m}",
                )
                t = st.selectbox(
                    f"Timing for month {m}",
                    ["beginning", "middle", "end"],
                    key=f"month_time_{m}",
                    index=1,
                )
                period_costs.append(c)
                period_timings.append(t)
        annual_om = st.number_input(
            "Annual O&M Cost ($)", min_value=0.0, value=0.0
        )
        annual_benefits = st.number_input(
            "Benefits (Annual, $)", min_value=0.0, value=0.0
        )
        base_year = st.number_input("Base Year (Year)", min_value=0, step=1, value=0)
        discount_rate = st.number_input(
            "Discount Rate (%)", min_value=0.0, value=0.0
        )
        period_analysis = st.number_input(
            "Period of Analysis (Years)", min_value=1, step=1, value=1
        )

        future_costs = []
        for i in range(st.session_state.num_future_costs):
            cost = st.number_input(
                f"Planned Future Cost {i + 1} ($)",
                min_value=0.0,
                value=0.0,
                key=f"fcost_{i}",
            )
            year = st.number_input(
                f"Year of Cost {i + 1}",
                min_value=0,
                step=1,
                value=st.session_state.get(f"fyear_{i}", base_year),
                key=f"fyear_{i}",
            )
            future_costs.append((cost, year))

        compute_annual = st.form_submit_button("Compute Annual Costs")

    if compute_annual:
        initial_cost = first_cost + real_estate_cost + ped_cost + monitoring_cost
        if idc_method == "Specify per-period costs":
            idc = interest_during_construction(
                initial_cost,
                idc_rate / 100.0,
                construction_months,
                costs=period_costs,
                timings=period_timings,
                normalize=False,
            )
        else:
            idc = interest_during_construction(
                initial_cost, idc_rate / 100.0, construction_months
            )
        total_initial = initial_cost + idc
        dr = discount_rate / 100.0
        future_details = []
        for cost, year in future_costs:
            pv_factor = 1 / ((1 + dr) ** (year - base_year))
            pv = cost * pv_factor
            future_details.append(
                {
                    "Cost": cost,
                    "Year": year,
                    "PV Factor": pv_factor,
                    "Present Value": pv,
                }
            )
        pv_future = sum(item["Present Value"] for item in future_details)
        total_investment = total_initial + pv_future
        crf = capital_recovery_factor(dr, period_analysis)
        annual_construction = total_investment * crf
        annual_total = annual_construction + annual_om
        bcr = annual_benefits / annual_total if annual_total else np.nan

        if future_details:
            future_df = pd.DataFrame(future_details)
            st.write("Planned Future Costs (Present Values)")
            st.table(future_df)
            st.success(f"Present Value of Future Costs: ${pv_future:,.2f}")
            st.session_state.future_costs_df = future_df
        else:
            st.session_state.future_costs_df = pd.DataFrame()

        st.success(f"Interest During Construction: ${idc:,.2f}")
        st.success(f"Total Cost/Investment: ${total_investment:,.2f}")
        st.success(f"Capital Recovery Factor: {crf:.4f}")
        st.success(
            f"Annual Construction Cost including O&M: ${annual_total:,.2f}"
        )
        st.success(f"Benefit-Cost Ratio: {bcr:,.2f}")

        st.session_state.annualizer_summary = {
            "Interest During Construction": idc,
            "Total Cost/Investment": total_investment,
            "Capital Recovery Factor": crf,
            "Annual Cost including O&M": annual_total,
            "Benefit-Cost Ratio": bcr,
        }
        st.session_state.annualizer_inputs = {
            "Project First Cost ($)": first_cost,
            "Real Estate Cost ($)": real_estate_cost,
            "PED Cost ($)": ped_cost,
            "Monitoring Cost ($)": monitoring_cost,
            "Interest Rate (%)": idc_rate,
            "Construction Period (Months)": construction_months,
            "IDC Cost Distribution": idc_method,
            "Annual O&M Cost ($)": annual_om,
            "Benefits (Annual, $)": annual_benefits,
            "Base Year": base_year,
            "Discount Rate (%)": discount_rate,
            "Period of Analysis (Years)": period_analysis,
        }

    export_button()


def udv_analysis():
    """Unit Day Value recreation benefit calculator."""
    st.header("Recreation Benefit (Unit Day Value)")
    st.info(
        "Estimate annual recreation benefits using USACE Unit Day Values (UDV).",
    )
    tab_calc, tab_rank = st.tabs(["Calculator", "Ranking Criteria"])
    with tab_calc:
        rec_type = st.selectbox(
            "Recreation Type",
            ["General", "Specialized"],
            help="Select the type of recreation experience.",
        )
        if rec_type == "General":
            activity = st.selectbox(
                "General Activity Type",
                ["General Recreation", "Fishing and Hunting"],
                help="Select the general recreation category.",
            )
        else:
            activity = st.selectbox(
                "Specialized Activity Type",
                ["Fishing and Hunting", "Other (e.g., Boating)"],
                help="Select the specialized recreation category.",
            )
        points = st.number_input(
            "Point Value",
            min_value=0.0,
            max_value=100.0,
            value=0.0,
            step=1.0,
            help="Total recreation ranking points (0-100).",
        )
        column_map = {
            ("General", "General Recreation"): "General Recreation",
            ("General", "Fishing and Hunting"): "General Fishing and Hunting",
            ("Specialized", "Fishing and Hunting"): "Specialized Fishing and Hunting",
            ("Specialized", "Other (e.g., Boating)"): "Specialized Recreation",
        }
        table_col = column_map[(rec_type, activity)]
        udv_calc = float(
            np.interp(
                points, POINT_VALUE_TABLE["Points"], POINT_VALUE_TABLE[table_col]
            )
        )
        udv_value = st.number_input(
            "Unit Day Value ($/user day)",
            min_value=0.0,
            value=udv_calc,
            help="Override if updated UDV schedules are available.",
        )
        user_days = st.number_input(
            "Expected Annual User Days",
            min_value=0.0,
            value=0.0,
            step=1.0,
        )
        visitation = st.number_input(
            "Expected Visitation",
            min_value=0.0,
            value=1.0,
            step=1.0,
            help="Multiplier applied to the expected annual user days.",
        )
        if st.button("Compute Recreation Benefit"):
            total_user_days = user_days * visitation
            benefit = udv_value * total_user_days
            st.success(f"Annual Recreation Benefit: ${benefit:,.2f}")
            st.info(f"Adjusted Annual User Days: {total_user_days:,.2f}")
            st.session_state.udv_benefit = benefit
            st.session_state.udv_inputs = {
                "Recreation Type": rec_type,
                "Activity Type": activity,
                "Point Value": points,
                "Unit Day Value": udv_value,
                "Expected Annual User Days": user_days,
                "Expected Visitation": visitation,
                "Adjusted Annual User Days": total_user_days,
            }
    with tab_rank:
        st.subheader(
            "Table 1. Guidelines for Assigning Points for General Recreation"
        )
        criteria_table = {
            "Criteria": [
                "Recreation experience",
                "Availability of opportunity",
                "Carrying capacity",
                "Accessibility",
                "Environmental quality",
            ],
            "Very Low": [
                "Two general activities (0-4)",
                "Several within 1 hr travel time; a few within 30 min (0-3)",
                "Minimum facility for public health and safety (0-2)",
                "Limited access by any means to site or within site (0-3)",
                "Low aesthetic quality; factors significantly lower quality (0-2)",
            ],
            "Low": [
                "Several general activities (5-10)",
                "Several within 1 hr travel time; none within 30 min (4-6)",
                "Basic facility to conduct activity(ies) (3-5)",
                "Fair access, poor quality roads to site; limited access within site (4-6)",
                "Average aesthetic quality; factors exist that lower quality (3-6)",
            ],
            "Moderate": [
                "Several general activities; one high quality value activity (11-16)",
                "One or two within 1 hr travel time; none within 30 min (7-10)",
                "Adequate facilities to conduct activity without resource deterioration (6-8)",
                "Fair access, fair road to site; fair access, good roads within site (7-10)",
                "Above average aesthetic quality; limiting factors can be rectified (7-10)",
            ],
            "High": [
                "Several general activities; more than one high quality value activity (17-23)",
                "None within 1 hr travel time; one or two within 2 hr travel time (11-14)",
                "Optimum facilities to conduct activity at site (9-11)",
                "Good access, good roads to site; fair access, good roads within site (11-14)",
                "High aesthetic quality; no factors exist that lower quality (11-15)",
            ],
            "Very High": [
                "Numerous high quality activities (24-30)",
                "None within 2 hr travel time (15-18)",
                "Ultimate potential facilities to achieve intent of selected alternative (12-14)",
                "Good access, high standard road to site; good access within site (15-18)",
                "Outstanding aesthetic quality; no factors exist that lower quality (16-20)",
            ],
        }
        st.table(pd.DataFrame(criteria_table).set_index("Criteria"))
        st.subheader("Table 2. Conversion of Points to Dollar Values")
        st.table(POINT_VALUE_TABLE.set_index("Points"))
    export_button()


def water_demand_forecast():
    """Municipal and industrial water demand forecast."""
    st.header("Water Demand Forecast")
    st.info(
        "Project municipal and industrial water demand using USACE guidance." \
        " Calculations follow ER 1105-2-100 methodology."
    )

    with st.form("water_demand_form"):
        base_year = st.number_input(
            "Base Year",
            min_value=0,
            value=2024,
            step=1,
            help="Starting year for the forecast.",
        )
        projection_years = st.number_input(
            "Projection Years",
            min_value=1,
            value=20,
            step=1,
            help="Number of years to project beyond the base year.",
        )
        base_pop = st.number_input(
            "Base Population",
            min_value=0.0,
            value=0.0,
            step=100.0,
            help="Population in the base year.",
        )
        per_capita = st.number_input(
            "Per-capita Municipal Demand (gallons/person/day)",
            min_value=0.0,
            value=100.0,
            help="Typical municipal use per person.",
        )

        pop_tab, ind_tab, loss_tab, cons_tab = st.tabs(
            ["Population", "Industry", "Losses", "Conservation"]
        )

        years = np.arange(base_year, base_year + projection_years + 1)
        growth_years = years[1:]

        with pop_tab:
            default_growth = st.number_input(
                "Default Annual Growth Rate (%)",
                value=1.0,
                step=0.1,
                help="Average annual population growth rate.",
            )
            growth_df = st.data_editor(
                pd.DataFrame(
                    {
                        "Year": growth_years,
                        "Growth Rate (%)": [default_growth] * len(growth_years),
                    }
                ),
                num_rows="dynamic",
                key="growth_df",
            )

        with ind_tab:
            default_ind = st.number_input(
                "Default Industrial Demand Factor (% of municipal)",
                min_value=0.0,
                value=20.0,
                step=1.0,
                help="Industrial demand as a percent of municipal demand.",
            )
            industrial_df = st.data_editor(
                pd.DataFrame(
                    {
                        "Year": years,
                        "Industrial Factor (%)": [default_ind] * len(years),
                    }
                ),
                num_rows="dynamic",
                key="industrial_df",
            )

        with loss_tab:
            default_loss = st.number_input(
                "Default System Losses (%)",
                min_value=0.0,
                value=10.0,
                step=1.0,
                help="Distribution losses as a percent of total demand.",
            )
            losses_df = st.data_editor(
                pd.DataFrame(
                    {
                        "Year": years,
                        "System Losses (%)": [default_loss] * len(years),
                    }
                ),
                num_rows="dynamic",
                key="losses_df",
            )

        with cons_tab:
            default_cons = st.number_input(
                "Default Conservation Reduction (%)",
                min_value=0.0,
                value=0.0,
                step=1.0,
                help="Percent reduction in demand from conservation measures.",
            )
            cons_df = st.data_editor(
                pd.DataFrame(
                    {
                        "Year": years,
                        "Conservation (%)": [default_cons] * len(years),
                    }
                ),
                num_rows="dynamic",
                key="cons_df",
            )

        submitted = st.form_submit_button("Run Forecast")

    if submitted:
        growth_rates = growth_df["Growth Rate (%)"].to_numpy() / 100
        pops = [base_pop]
        for gr in growth_rates:
            pops.append(pops[-1] * (1 + gr))
        pops = np.array(pops)

        industrial_factors = industrial_df["Industrial Factor (%)"].to_numpy() / 100
        loss_factors = losses_df["System Losses (%)"].to_numpy() / 100
        conservation_rates = cons_df["Conservation (%)"].to_numpy() / 100
        per_capita_yearly = per_capita * (1 - conservation_rates)

        municipal_mgy = pops * per_capita_yearly * 365 / 1e6
        industrial_mgy = municipal_mgy * industrial_factors
        total_mgy = (municipal_mgy + industrial_mgy) * (1 + loss_factors)

        input_df = pd.DataFrame(
            {
                "Year": years,
                "Population": pops.round(0).astype(int),
                "Growth Rate (%)": np.concatenate(([np.nan], growth_rates * 100)),
                "Industrial Factor (%)": industrial_factors * 100,
                "System Losses (%)": loss_factors * 100,
                "Conservation (%)": conservation_rates * 100,
            }
        )
        result_df = pd.DataFrame(
            {
                "Year": years,
                "Population": pops.round(0).astype(int),
                "Municipal Demand (MGY)": municipal_mgy,
                "Industrial Demand (MGY)": industrial_mgy,
                "Total Demand (MGY)": total_mgy,
            }
        )
        st.session_state.water_input_table = input_df
        st.session_state.water_demand_results = result_df

    if st.session_state.get("water_input_table") is not None:
        st.subheader("Input Assumptions")
        st.table(st.session_state.water_input_table)

    result_df = st.session_state.get("water_demand_results", pd.DataFrame())
    if not result_df.empty:
        st.subheader("Forecast Results")
        st.table(result_df)
        st.line_chart(result_df.set_index("Year")["Total Demand (MGY)"])

    export_button()


def readme_page():
    """Display repository README."""
    st.header("ReadMe")
    st.markdown(Path("README.md").read_text())


section = st.sidebar.radio(
    "Navigate",
    [
        "EAD Calculator",
        "Updated Storage Cost",
        "Project Annualizer",
        "UDV Analysis",
        "Water Demand Forecast",
        "ReadMe",
    ],
)

if section == "EAD Calculator":
    ead_calculator()
elif section == "Updated Storage Cost":
    storage_calculator()
elif section == "Project Annualizer":
    annualizer_calculator()
elif section == "UDV Analysis":
    udv_analysis()
elif section == "Water Demand Forecast":
    water_demand_forecast()
else:
    readme_page()

